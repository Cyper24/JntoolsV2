import streamlit as st
import time
import os
import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import subprocess
from docxtpl import DocxTemplate
from datetime import datetime
import locale

at=(st.session_state.ato).strip(" ")
st.title("Manifest")

st.caption("Input Kode Tugas")
kt = st.text_input("")

if st.button("Exec..."):
    st.text("Waitt...")
    list = []
    headers = {
                "cookie": "HWWAFSESID=a00e27f02785ef49ce5; HWWAFSESTIME=1738201375713",
                "authtoken": f"{at}",
                "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/132.0.0.0 Safari/537.36 Edg/132.0.0.0",
                "Content-Type": "application/json"
    }
    urlkt = "https://jmsgw.jntexpress.id/transportation/trackingDeatil/loading/scan/page"
    urlkt2 = "https://jmsgw.jntexpress.id/transportation/tmsShipment/traceDetail"
    querystring = {"current":"1","size":"20000","shipmentNo":f"{kt}","scanNetworkCode":"SOC999"}
    querystring2 = {"shipmentNo":f"{kt}"}
    response = requests.request("GET", urlkt, headers=headers, params=querystring)
    response2 = requests.request("GET", urlkt2, headers=headers, params=querystring2)
    manifest = response.json()
    manifest2 = response2.json()
    fdocx = manifest2["data"]["shipmentDetail"]
    tujuan = fdocx["endName"]
    nopol = fdocx["plateNumber"]
    driver = fdocx["driverName"]
    tgld = fdocx["plannedDepartureTime"]
    locale.setlocale(locale.LC_TIME, 'id_ID')
    xt = datetime.strptime(f'{tgld}', '%Y-%m-%d %H:%M:%S')
    tgal = xt.strftime("%A, %d-%m-%Y / %H.%M")
    doc = DocxTemplate("sj.docx")
    context = {'driver' : driver,
           'nopol': nopol,
            'kt' : kt,
            'tujuan' : tujuan,
            'tgl' : tgal}
    doc.render(context)
    doc.save("sjnew.docx")

    f = manifest["data"]["records"]
    for x in f:
        billCode = x["billCode"]
        packageCode = x["packageCode"]
        final = {'No. Waybill' : billCode,'Kepemilikan No. Bagging' : packageCode}
        list.append(final)
    df = pd.DataFrame(list)
    df.to_excel('temp.xlsx')
    df = pd.read_excel('temp.xlsx')
    df['Kepemilikan No. Bagging'] = df['Kepemilikan No. Bagging'].fillna("-" + kt)
    pivot = df.pivot_table(index=["Kepemilikan No. Bagging"],values=['No. Waybill'],aggfunc=['count'],margins=True, margins_name='Total')
    rows2 = pivot["count"]["No. Waybill"].reset_index()
    wb = Workbook()
    sheet = wb.active
    for r in dataframe_to_rows(rows2):
            sheet.append(r)
        # sheet.delete_cols(idx=1)
    sheet.delete_rows(idx=2)
    sheet.insert_rows(idx=1,amount=2)
    sheet["A1"] = "SOC GATEWAY"
    sheet["A2"] = "OUTGOING SOC GATEWAY TO " + tujuan
    sheet["B3"] = "No Bagging "
    sheet["C3"] = "AWB"
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)
    tc = int(len(sheet['A']))
    bagging = tc - 5
    ecer = sheet["C4"].value - bagging
    sheet["C4"] = ecer
    tr = int(sheet[f"C{tc}"].value)
    total = tr - bagging
    sheet[f"C{tc}"] = total
    bord = Border(left=Side(style='thin'), 
                           right=Side(style='thin'), 
                           top=Side(style='thin'), 
                           bottom=Side(style='thin'))
    for row in range(1,sheet.max_row+1):
                for col in range(1,sheet.max_column+1):
                        cell=sheet.cell(row, col)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.font = Font(color="000000", size=11,bold=True)
                        cell.border = bord
    bold= Font(color="000000", size=14,bold=True)
    sheet["A1"].font = bold
    sheet["A2"].font = bold
    wb.save(filename="manifest.xlsx")
    st.text("Done...")

    subprocess.Popen(["manifest.xlsx"],shell=True)
    time.sleep(0.5)
    subprocess.Popen(["sjnew.docx"],shell=True) 
